import telegram
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
import telebot
from openpyxl import load_workbook
import datetime



fn = 'C:/Users/mcpav/PycharmProjects/pythonProject_Botyra/Summer_bot/o.xlsx'
wb = load_workbook(fn)
ws = wb['data']
count = 1
fr = 'C:/Users/mcpav/PycharmProjects/pythonProject_Botyra/Summer_bot/ref.xlsx'
wt = load_workbook(fr)
wn = wt['data']
counter = 1

STATE = None


def start(update, context):
    first_name = update.message.chat.first_name
    update.message.reply_text(f"{first_name}, здравствуйте!")


def review(update, context):
    global STATE
    update.message.reply_text("Оставьте отзыв или предложение, мы обязательно на него отреагируем")
    STATE = 2


def reviewer(update, context):
    global counter
    recyve = update.message.text
    context.user_data['sos'] = recyve
    update.message.reply_text(f"Спасибо за ваш отзыв")
    first_name = update.message.chat.first_name
    chat_id = update.update_id
    user_name = update.message.chat.username
    wn[f'A{counter}'] = str(chat_id)
    wn[f'B{counter}'] = user_name
    wn[f'C{counter}'] = first_name
    wn[f'D{counter}'] = recyve
    counter += 1
    wt.save(fr)
    wt.close()


def map(update, context):
    doc_file = open('toyota.jpg', 'rb')
    chat_id = update.effective_chat.id
    context.bot.send_document(chat_id, doc_file)

def sos(update, context):
    global STATE
    first_name = update.message.chat.first_name
    update.message.reply_text(f"{first_name}, для вызова помощи напишите: адрес, ваше имя, ваш номер телефона, кого нужно вызвать(полицию/пожарных/скорую)")
    STATE = 1

def sos_rec(update, context):
    global count
    recyve = update.message.text
    context.user_data['sos'] = recyve
    mas = recyve.split(",")
    update.message.reply_text(f"Помощь вызвана, ожидайте")
    ws[f'A{count}'] = str(datetime.datetime.now())[:19]
    ws[f'B{count}'] = mas[0]
    ws[f'C{count}'] = mas[1]
    ws[f'D{count}'] = mas[2]
    ws[f'E{count}'] = mas[3]
    count+=1
    wb.save(fn)
    wb.close()


def help(update, context):
    update.message.reply_text('получена команда помощи')


def error(update, context):
    update.message.reply_text('Произошла ошибка')


def text(update, context):
    global STATE
    if STATE == 1:
        return sos_rec(update,context)
    elif STATE == 2:
        return reviewer(update,context)


def main():
    TOKEN = "5216047631:AAE6UKsIbuiB9NSblWWUvoB_K4x_FB5oPpI"
    updater = Updater(TOKEN, use_context=True)
    dispatcher = updater.dispatcher
    dispatcher.add_handler(CommandHandler("start", start))
    dispatcher.add_handler(CommandHandler("map", map))
    dispatcher.add_handler(CommandHandler("review", review))
    dispatcher.add_handler(CommandHandler("sos", sos))
    dispatcher.add_handler(CommandHandler("help", help))
    dispatcher.add_handler(MessageHandler(Filters.text,text))
    dispatcher.add_error_handler(error)
    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    main()